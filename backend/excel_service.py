import pandas as pd
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

class ExcelService:
    """Service for parsing and generating Excel files"""
    
    @staticmethod
    def parse_upload(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
        """Parse uploaded Excel file and return list of items"""
        try:
            # Read Excel file
            df = pd.read_excel(BytesIO(file_content))
            
            # Normalize column names (remove spaces, lowercase)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            # Expected columns: refid, medida, marca, modelo, indice, meu_preco
            # Or variations: ref_id, meupreco, etc.
            
            # Map common variations
            column_mapping = {
                'ref_id': 'refid',
                'ref': 'refid',
                'id': 'refid',
                'meupreco': 'meu_preco',
                'meu_custo': 'meu_preco',
                'preco': 'meu_preco',
                'custo': 'meu_preco',
                'indicevelocidadecarga': 'indice',
                'indice_velocidade_carga': 'indice',
                'xl': 'indice',
            }
            
            # Apply mapping
            df = df.rename(columns=column_mapping)
            
            # Validate required columns
            required_cols = ['medida', 'marca', 'modelo', 'indice', 'meu_preco']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                raise ValueError(f"Missing required columns: {', '.join(missing_cols)}")
            
            # Add RefID if not present
            if 'refid' not in df.columns:
                df['refid'] = range(1, len(df) + 1)
            
            # Convert to list of dicts
            items = []
            for idx, row in df.iterrows():
                try:
                    item = {
                        'ref_id': str(row['refid']),
                        'medida': str(row['medida']).strip(),
                        'marca': str(row['marca']).strip(),
                        'modelo': str(row['modelo']).strip(),
                        'indice': str(row['indice']).strip() if pd.notna(row.get('indice')) else '',
                        'meu_preco': float(row['meu_preco']),
                    }
                    items.append(item)
                except Exception as e:
                    logger.warning(f"Skipping row {idx}: {str(e)}")
                    continue
            
            logger.info(f"Parsed {len(items)} items from {filename}")
            return items
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise ValueError(f"Failed to parse Excel file: {str(e)}")
    
    @staticmethod
    def generate_results(job: Dict[str, Any], items: List[Dict[str, Any]], 
                         suppliers: List[str]) -> bytes:
        """Generate results Excel file with formatting"""
        try:
            # Prepare data for DataFrame
            data = []
            for item in items:
                row = {
                    'RefID': item['ref_id'],
                    'Medida': item['medida'],
                    'Marca': item['marca'],
                    'Modelo': item['modelo'],
                    'Indice': item['indice'],
                    'MeuPreço': item['meu_preco'],
                    'MelhorPreço': item.get('melhor_preco', ''),
                    'Fornecedor': item.get('melhor_fornecedor', ''),
                    'Economia€': item.get('economia_euro', ''),
                    'Economia%': item.get('economia_percent', ''),
                }
                
                # Add supplier columns
                supplier_prices = item.get('supplier_prices', {})
                for supplier_name in suppliers:
                    price = supplier_prices.get(supplier_name, '')
                    if isinstance(price, (int, float)):
                        row[supplier_name] = price
                    else:
                        row[supplier_name] = price if price else 'NAO_ENCONTRADO'
                
                row['Status'] = item.get('status', 'pending').upper()
                data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Write to Excel with formatting
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resultados', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Resultados']
                
                # Define styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                
                savings_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")  # Light green
                
                border = Border(
                    left=Side(style='thin', color='d1d5db'),
                    right=Side(style='thin', color='d1d5db'),
                    top=Side(style='thin', color='d1d5db'),
                    bottom=Side(style='thin', color='d1d5db')
                )
                
                # Format header row
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Find Economia€ column index (1-based) dynamically
                economia_col_idx = next(
                    (i + 1 for i, col in enumerate(df.columns) if col == 'Economia€'),
                    9,
                )

                # Format data rows
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')

                        # Highlight rows with savings
                        if cell.column == economia_col_idx:
                            if isinstance(cell.value, (int, float)) and cell.value > 0:
                                for c in row:
                                    c.fill = savings_fill
                
                # Adjust column widths
                column_widths = {
                    'A': 10,  # RefID
                    'B': 15,  # Medida
                    'C': 15,  # Marca
                    'D': 20,  # Modelo
                    'E': 12,  # Indice
                    'F': 12,  # MeuPreço
                    'G': 12,  # MelhorPreço
                    'H': 20,  # Fornecedor
                    'I': 12,  # Economia€
                    'J': 12,  # Economia%
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Set remaining columns (suppliers) to width 12
                for col_idx in range(11, worksheet.max_column + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 15
            
            output.seek(0)
            return output.read()
            
        except Exception as e:
            logger.error(f"Error generating Excel results: {str(e)}")
            raise ValueError(f"Failed to generate Excel results: {str(e)}")
